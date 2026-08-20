from __future__ import annotations

import argparse
import importlib.util
from pathlib import Path
import sys

from openpyxl import load_workbook


def load_module(project_root: Path):
    module_path = project_root / "关键词预警资料迭代_单文件版.py"
    spec = importlib.util.spec_from_file_location("keyword_warning_evolution_mock", module_path)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--prompt", required=True)
    cli = parser.parse_args()

    project_root = Path(__file__).resolve().parents[1]
    evo = load_module(project_root)
    workbook = load_workbook(cli.input, read_only=False, data_only=False)
    sheet = evo.data_sheet(workbook)
    max_row, max_col = evo.used_bounds(sheet)
    headers = [evo.text(sheet.cell(1, col).value) for col in range(1, max_col + 1)]
    col_map = {name: index + 1 for index, name in enumerate(headers) if name}
    tasks = evo.collect_tasks(sheet, col_map, max_row, 0)
    results = evo.collect_results(sheet, col_map, tasks)

    args = argparse.Namespace(
        no_badcase_export=False,
        knowledge_dir=str(project_root / "knowledge"),
        gap_max_knowledge_chars=60000,
        gap_model="mock-analysis-model",
        model="mock-warning-model",
        gap_thinking="off",
        gap_min_support=2,
        analyze_gaps=True,
        gap_force=True,
        gap_batch_size=15,
        gap_max_dialogue_chars=4000,
    )

    def fake_structured_model(_args, system_prompt, payload, purpose):
        if "BadCase归因" in purpose:
            results_payload = []
            for item in payload["badcases"]:
                is_false_positive = "误预警" in item["case_types"]
                results_payload.append(
                    {
                        "badcase_id": item["badcase_id"],
                        "attribution": "标签边界缺口" if is_false_positive else "资料缺口",
                        "confidence": "高",
                        "gap_type": "Hard Negative" if is_false_positive else "新表达",
                        "evidence": "离线 Mock：现有示例资料未覆盖该模式。",
                        "knowledge_coverage": "离线 Mock，待人工核对正式资料。",
                        "common_scene": "用户表达公开传播意图，或普通咨询被误判。",
                        "candidate_knowledge": "离线 Mock 候选内容，不可直接用于生产。",
                        "risk": "可能把普通投诉或咨询误判为目标标签。",
                        "action": "补边界" if is_false_positive else "补资料",
                    }
                )
            return {"results": results_payload}, "{}"

        if "Knowledge Gap聚类" in purpose:
            ids = [item["badcase_id"] for item in payload["badcases"]]
            return (
                {
                    "knowledge_gaps": [
                        {
                            "target_labels": ["B1"],
                            "gap_type": "标签边界",
                            "title": "公开传播意图与普通咨询的边界",
                            "description": "离线 Mock：需要同时补正向证据和排除条件。",
                            "support_case_ids": ids,
                            "common_pattern": "存在公开传播表达，或普通咨询被错误联想。",
                            "suggested_rule": "离线 Mock 候选：必须出现明确传播对象、渠道或行为计划；仅查询订单不命中。",
                            "positive_evidence": "明确提出公开、发布、曝光或向特定公众渠道传播。",
                            "negative_boundary": "仅投诉、咨询、表达不满但没有传播意图时不命中。",
                            "minimum_context_chain": "无",
                            "risk": "‘投诉’与‘公开传播’可能被过度合并。",
                            "conflict_labels": ["A4"],
                            "recommended_action": "补边界",
                        }
                    ]
                },
                "{}",
            )
        raise AssertionError(f"unexpected purpose: {purpose}; prompt={system_prompt[:20]}")

    original = evo.call_structured_model
    evo.call_structured_model = fake_structured_model
    try:
        evo.run_knowledge_evolution_pipeline(
            workbook,
            sheet,
            col_map,
            tasks,
            results,
            args,
            Path(cli.prompt).read_text(encoding="utf-8"),
            Path(cli.output),
        )
    finally:
        evo.call_structured_model = original
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

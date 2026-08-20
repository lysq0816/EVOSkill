from __future__ import annotations

import importlib.util
from pathlib import Path
import sys
from types import SimpleNamespace
import unittest

from openpyxl import Workbook, load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
MODULE_PATH = PROJECT_ROOT / "关键词预警资料迭代_单文件版.py"
SPEC = importlib.util.spec_from_file_location("keyword_warning_evolution", MODULE_PATH)
assert SPEC is not None and SPEC.loader is not None
evo = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = evo
SPEC.loader.exec_module(evo)


class EvolutionPipelineTests(unittest.TestCase):
    def setUp(self) -> None:
        evo.LABEL_ALIASES.clear()
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "数据"
        self.headers = [
            "keyid",
            "业务线",
            evo.INPUT_COL,
            evo.TRUE_LABEL_COL,
            evo.PRED_WARN_COL,
            evo.PRED_LABEL_COL,
            evo.ELAPSED_COL,
            evo.REASON_COL,
            "模型原始输出",
            "请求错误",
        ]
        self.sheet.append(self.headers)
        rows = [
            ["k1", "酒店", "用户：我要曝光这件事", "B1"],
            ["k2", "酒店", "用户：正常咨询订单", ""],
            ["k3", "机票", "用户：客服服务差并准备公开投诉", "A4、B1"],
            ["k4", "火车", "用户：有点失望", "【3-4】"],
            ["k5", "酒店", "用户：仍未解决", "B0"],
        ]
        for row in rows:
            self.sheet.append(row)
        self.col_map = {name: index + 1 for index, name in enumerate(self.headers)}
        self.tasks = [
            evo.RowTask(row_num=row_num, index=row_num - 1, dialogue=row[2], true_label=row[3])
            for row_num, row in enumerate(rows, start=2)
        ]
        self.results = {
            2: evo.EvalResult(2, "否", "", "未命中", 0.1, "{}"),
            3: evo.EvalResult(3, "是", "B1", "判断为舆情", 0.1, "{}"),
            4: evo.EvalResult(4, "是", "A4、B0", "标签混淆", 0.1, "{}"),
            5: evo.EvalResult(5, "否", "【3-4】", "轻度负面", 0.1, "{}"),
            6: evo.EvalResult(6, "", "", "请求失败", 0.1, "", error="timeout"),
        }

    def test_extract_badcases_classifies_warning_and_label_errors(self) -> None:
        badcases = evo.extract_badcases(
            self.sheet,
            self.col_map,
            self.tasks,
            self.results,
        )
        self.assertEqual(4, len(badcases))
        by_row = {case.row_num: case for case in badcases}
        self.assertEqual(["漏预警", "标签漏召"], by_row[2].case_types)
        self.assertEqual(["B1"], by_row[2].missing_labels)
        self.assertEqual(["误预警", "标签误报"], by_row[3].case_types)
        self.assertEqual(["标签混淆"], by_row[4].case_types)
        self.assertEqual(["B1"], by_row[4].missing_labels)
        self.assertEqual(["B0"], by_row[4].extra_labels)
        self.assertEqual(["请求错误"], by_row[6].case_types)
        self.assertNotIn(5, by_row)

    def test_manual_review_survives_badcase_sheet_rebuild(self) -> None:
        badcases = evo.extract_badcases(
            self.sheet,
            self.col_map,
            self.tasks,
            self.results,
        )
        analyses = {
            case.badcase_id: evo.default_badcase_analysis(case, "GE-test")
            for case in badcases
        }
        evo.write_badcase_sheet(self.workbook, badcases, analyses, {})
        ws = self.workbook[evo.BADCASE_SHEET]
        header_map = {
            ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)
        }
        ws.cell(2, header_map["人工归因"], "资料缺口")
        ws.cell(2, header_map["人工备注"], "已由人工确认")
        existing = evo.read_sheet_records(
            self.workbook,
            evo.BADCASE_SHEET,
            "BadCase ID",
        )
        evo.write_badcase_sheet(self.workbook, badcases, analyses, existing)
        rebuilt = self.workbook[evo.BADCASE_SHEET]
        rebuilt_headers = {
            rebuilt.cell(1, col).value: col for col in range(1, rebuilt.max_column + 1)
        }
        self.assertEqual("资料缺口", rebuilt.cell(2, rebuilt_headers["人工归因"]).value)
        self.assertEqual("已由人工确认", rebuilt.cell(2, rebuilt_headers["人工备注"]).value)

    def test_gap_cluster_recalculates_support_and_governance(self) -> None:
        badcases = evo.extract_badcases(
            self.sheet,
            self.col_map,
            self.tasks,
            self.results,
        )
        case_by_id = {case.badcase_id: case for case in badcases}
        ids = list(case_by_id)
        item = {
            "target_labels": ["B1"],
            "gap_type": "新表达",
            "title": "公开传播意图",
            "description": "资料缺少间接表达",
            "support_case_ids": [ids[0], ids[1], "BC-unknown"],
            "suggested_rule": "当用户明确表示将问题公开传播时，结合业务定义判断 B1。",
            "positive_evidence": "明确传播对象或渠道",
            "negative_boundary": "仅表达不满但没有传播意图时不命中",
            "risk": "可能与普通投诉混淆",
            "conflict_labels": [],
        }
        cluster = evo.normalize_gap_cluster(
            item,
            set(case_by_id),
            case_by_id,
            min_support=2,
        )
        self.assertIsNotNone(cluster)
        assert cluster is not None
        self.assertEqual(2, len(cluster.support_case_ids))
        self.assertEqual("待人工审核", cluster.governance_status)
        self.assertTrue(cluster.cluster_id.startswith("KG-"))

    def test_short_knowledge_context_is_not_marked_truncated(self) -> None:
        context, metadata = evo.read_knowledge_context(
            "## B1\n示例资料\n",
            PROJECT_ROOT / "tests" / "does-not-exist",
            max_chars=1000,
        )
        self.assertFalse(metadata["truncated"])
        self.assertEqual(metadata["original_chars"], metadata["sent_chars"])
        self.assertIn("示例资料", context)

    def test_attribution_batch_is_normalized_and_checkpointed(self) -> None:
        badcases = evo.extract_badcases(
            self.sheet,
            self.col_map,
            self.tasks,
            self.results,
        )
        analyses = {
            case.badcase_id: evo.default_badcase_analysis(case, "GE-test")
            for case in badcases
        }
        checkpoints: list[int] = []

        def fake_call(_args, _system_prompt, payload, purpose):
            self.assertIn("BadCase归因", purpose)
            return (
                {
                    "results": [
                        {
                            "badcase_id": item["badcase_id"],
                            "attribution": "knowledge_gap",
                            "confidence": "high",
                            "gap_type": "新表达",
                            "evidence": "现有资料未覆盖",
                            "candidate_knowledge": "候选规则",
                            "action": "补资料",
                        }
                        for item in payload["badcases"]
                    ]
                },
                "{}",
            )

        original = evo.call_structured_model
        evo.call_structured_model = fake_call
        try:
            evo.attribute_badcases(
                SimpleNamespace(gap_batch_size=2, gap_max_dialogue_chars=1000),
                SimpleNamespace(),
                "当前资料",
                badcases,
                analyses,
                "GE-test",
                checkpoint=lambda: checkpoints.append(1),
            )
        finally:
            evo.call_structured_model = original

        analyzed = [
            item for item in analyses.values() if item["analysis_status"] == "已完成"
        ]
        self.assertEqual(3, len(analyzed))
        self.assertTrue(all(item["attribution"] == "资料缺口" for item in analyzed))
        self.assertEqual(2, len(checkpoints))

    def test_pipeline_without_model_exports_review_sheets(self) -> None:
        args = SimpleNamespace(
            no_badcase_export=False,
            knowledge_dir=str(PROJECT_ROOT / "knowledge"),
            gap_max_knowledge_chars=60000,
            gap_model="",
            model="test-model",
            gap_thinking="off",
            gap_min_support=2,
            analyze_gaps=False,
            gap_force=False,
        )
        output_path = PROJECT_ROOT / "tests" / "tmp_pipeline_output.xlsx"
        try:
            summary = evo.run_knowledge_evolution_pipeline(
                self.workbook,
                self.sheet,
                self.col_map,
                self.tasks,
                self.results,
                args,
                "## B1【舆情】\n测试资料",
                output_path,
            )
            self.assertEqual(4, summary["badcases"])
            saved = load_workbook(output_path, read_only=False, data_only=False)
            for sheet_name in (
                evo.BADCASE_SHEET,
                evo.GAP_SHEET,
                evo.CANDIDATE_SHEET,
                evo.EVO_SUMMARY_SHEET,
            ):
                self.assertIn(sheet_name, saved.sheetnames)
            self.assertEqual(5, saved[evo.BADCASE_SHEET].max_row)
        finally:
            output_path.unlink(missing_ok=True)


if __name__ == "__main__":
    unittest.main()
